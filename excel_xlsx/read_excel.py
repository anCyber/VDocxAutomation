import openpyxl
import os

from excel_xlsx.popup_windows import (
    ___AttemptedToReadTooManyRows_PopUpWindow___,
    ___AttemptedToReadNegOrZeroRowsOrIncludesSymbols_PopUpWindow___
)

CONVERT_LETTERS_TO_NUMBERS = 64
ADD_TO_COLUMNS_AMOUNT = 6

def get_maximum_table_rows(worksheet: openpyxl.worksheet, first_cell_row: int, first_cell_column_letter: str) -> int:
    ''' Calculates the maximum amount of rows in the table, that contain data about employees '''
    first_cell_column = ord(f"{first_cell_column_letter}") - CONVERT_LETTERS_TO_NUMBERS
    READ_COLUMNS_AMOUNT = 5 + first_cell_column

    amount_of_rows: int = 0
    for table_row in worksheet.iter_rows(first_cell_row, worksheet.max_row, first_cell_column, READ_COLUMNS_AMOUNT):
        is_row_empty: bool = True
        for cell in table_row:
            if (cell.value is not None):
                is_row_empty = False

        if (not is_row_empty):
            amount_of_rows += 1
        
    return amount_of_rows


def convert_datetime_object_to_str(datetime_object) -> str:
    ''' Converts the datetime object into formatted string. Strings Format: "%d.%m.%Y" '''
    datetime_object_string = str(datetime_object)
    datetime_object_string = datetime_object_string[:10]
    separated_YMD_date_string = datetime_object_string.split("-")
    formated_DMY_string = f"{separated_YMD_date_string[2]}.{separated_YMD_date_string[1]}.{separated_YMD_date_string[0]}"
    return formated_DMY_string


def iterate_through_table_rows(worksheet: openpyxl.worksheet, read_rows_amount: int, first_cell_row: int, first_cell_column_letter: str):
    ''' Iterates through table rows and returns the list of rows, containing desired information '''
    first_cell_column = ord(f"{first_cell_column_letter}") - CONVERT_LETTERS_TO_NUMBERS
    READ_COLUMNS_AMOUNT = 5 + first_cell_column 
    try:
        list_of_rows: list = []
        for table_row in worksheet.iter_rows(first_cell_row, (read_rows_amount + first_cell_row - 1), first_cell_column, READ_COLUMNS_AMOUNT):
            table_row_list = []
            for cell in table_row:
                table_row_list.append(cell.value)
            
            table_row_list[-1] = convert_datetime_object_to_str(table_row_list[-1])
            table_row_list.pop(3)  # Removing undesired column from list (tabel_number)
            list_of_rows.append(table_row_list)
        maximum_rows = get_maximum_table_rows(worksheet, first_cell_row, first_cell_column_letter)

        if (maximum_rows >= read_rows_amount*2):
            worksheet.move_range(f"{first_cell_column_letter}{first_cell_row + read_rows_amount}:{chr(ord(f"{first_cell_column_letter}") + ADD_TO_COLUMNS_AMOUNT)}{maximum_rows + first_cell_row - 1}",
                                rows=-read_rows_amount, cols=0)
        else:
            if (maximum_rows <= read_rows_amount):
                worksheet.delete_rows(first_cell_row, first_cell_row + maximum_rows)
            else:
                worksheet.move_range(f"{first_cell_column_letter}{first_cell_row + read_rows_amount}:{chr(ord(f"{first_cell_column_letter}") + ADD_TO_COLUMNS_AMOUNT)}{maximum_rows + first_cell_row - 1}",
                                    rows=-read_rows_amount, cols=0)
                worksheet.delete_rows(first_cell_row + (maximum_rows - read_rows_amount), first_cell_row + maximum_rows)      
        
    except IndexError:
        list_of_rows: list = []
        maximum_rows = get_maximum_table_rows(worksheet, first_cell_row, first_cell_column_letter)
        ___AttemptedToReadTooManyRows_PopUpWindow___(maximum_rows)
        read_rows_amount = maximum_rows
        for table_row in worksheet.iter_rows(first_cell_row, (read_rows_amount + first_cell_row - 1), first_cell_column, READ_COLUMNS_AMOUNT):
            table_row_list = []
            for cell in table_row:
                table_row_list.append(cell.value)
            
            table_row_list[-1] = convert_datetime_object_to_str(table_row_list[-1])
            table_row_list.pop(3)  # Removing undesired column from list (tabel_number)
            list_of_rows.append(table_row_list)
        worksheet.delete_rows(first_cell_row, first_cell_row + maximum_rows)

    return list_of_rows



def get_list_of_info_about_employees(PATH_TO_DATA_FILE: str, read_rows_amount: int, first_cell_row: int, first_cell_column_letter: str, should_create_copy_of_excel: bool):
    '''
        Reads the Excel-file and returns extracted Dataframe. Contains following columns:
        "Структурное подразделение", "Должность (специальность, профессия) по штатному расписанию", "Фамилия, имя отчество", "количество календарных дней", "запланированная"
    '''
    if (read_rows_amount <= 0):
        ___AttemptedToReadNegOrZeroRowsOrIncludesSymbols_PopUpWindow___()
        
    else:
        if (should_create_copy_of_excel):
            excel_file_copy = openpyxl.load_workbook(PATH_TO_DATA_FILE)
            path_to_copy = f"{PATH_TO_DATA_FILE.removesuffix(".xlsx")}_BackUpCopy"
            while (os.path.exists(os.path.join(os.getcwd(), f"{path_to_copy}.xlsx"))):
                path_to_copy += "Newer"
            excel_file_copy.save(f"{path_to_copy}.xlsx")

        excel_file = openpyxl.load_workbook(PATH_TO_DATA_FILE)
        first_sheet = excel_file.worksheets[0]
        list_of_table_rows = iterate_through_table_rows(first_sheet, read_rows_amount, first_cell_row, first_cell_column_letter)
        excel_file.save(PATH_TO_DATA_FILE)
        return list_of_table_rows 